import io
from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.daily_task import DailyTask, DailyTaskLabel, DailyTaskStatus
from app.models.user import User

router = APIRouter(prefix="/daily-tasks", tags=["daily-tasks"])

# Excel 欄位定義（欄位順序 / 名稱 / 必填）
_COLUMNS = [
    ("標題", "title", True),
    ("日期", "date", True),  # yyyy-mm-dd
    ("狀態", "status", False),  # pending/in_progress/done/cancelled
    ("進度%", "progress", False),  # 0-100
    ("標籤", "labels", False),  # 逗號分隔
    ("說明", "description", False),
    ("工作分鐘數", "work_minutes", False),  # 整數
]

_STATUS_MAP = {
    "待辦": "pending",
    "pending": "pending",
    "進行中": "in_progress",
    "in_progress": "in_progress",
    "完成": "done",
    "done": "done",
    "已取消": "cancelled",
    "cancelled": "cancelled",
}


def _parse_row(row: tuple[Any, ...], row_idx: int) -> dict | str:
    """Parse one Excel row. Returns dict on success, error string on failure."""
    cells = [c.value for c in row]
    # pad to 7 columns
    while len(cells) < 7:
        cells.append(None)

    title = str(cells[0]).strip() if cells[0] is not None else ""
    if not title:
        return f"第 {row_idx} 列：標題不可為空"

    # 日期解析
    raw_date = cells[1]
    parsed_date: date | None = None
    if isinstance(raw_date, datetime):
        parsed_date = raw_date.date()
    elif isinstance(raw_date, date):
        parsed_date = raw_date
    elif isinstance(raw_date, str):
        try:
            parsed_date = date.fromisoformat(raw_date.strip())
        except ValueError:
            return f"第 {row_idx} 列：日期格式錯誤（應為 yyyy-mm-dd），得到「{raw_date}」"
    else:
        return f"第 {row_idx} 列：日期欄位為空"

    # 狀態
    raw_status = str(cells[2]).strip() if cells[2] is not None else "pending"
    status = _STATUS_MAP.get(raw_status, "pending")

    # 進度
    try:
        progress = int(cells[3]) if cells[3] is not None else 0
        progress = max(0, min(100, progress))
    except (TypeError, ValueError):
        progress = 0

    # 標籤
    raw_labels = str(cells[4]).strip() if cells[4] is not None else ""
    labels = [lbl.strip() for lbl in raw_labels.split(",") if lbl.strip()] if raw_labels else []

    # 說明
    description = str(cells[5]).strip() if cells[5] is not None else None
    if not description:
        description = None

    # 工作分鐘數
    try:
        work_minutes = int(cells[6]) if cells[6] is not None else 0
        work_minutes = max(0, work_minutes)
    except (TypeError, ValueError):
        work_minutes = 0

    return {
        "title": title,
        "date": parsed_date,
        "status": status,
        "progress": progress,
        "labels": labels,
        "description": description,
        "work_minutes": work_minutes,
    }


@router.get("/import/template")
async def download_template(_: User = Depends(get_current_user)):
    """下載 Excel 匯入範本"""
    wb = Workbook()
    ws = wb.active
    ws.title = "日常作業"

    # 標題列樣式
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [col[0] for col in _COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 欄寬
    col_widths = [20, 14, 12, 10, 20, 30, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # 範例資料
    examples = [
        ["每日站立會議", "2026-06-03", "完成", 100, "開發, 會議", "15 分鐘快速同步", 15],
        ["撰寫技術文件", "2026-06-03", "進行中", 50, "文件", "API 文件更新", 60],
        ["Code Review", "2026-06-04", "待辦", 0, "開發", "", 0],
    ]
    for r_idx, row in enumerate(examples, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # 備註頁
    ws2 = wb.create_sheet("填寫說明")
    notes = [
        ["欄位", "必填", "格式說明"],
        ["標題", "是", "任意文字，最多 500 字"],
        ["日期", "是", "yyyy-mm-dd，例：2026-06-03"],
        ["狀態", "否", "待辦 / 進行中 / 完成 / 已取消（預設：待辦）"],
        ["進度%", "否", "0 ~ 100 整數（預設：0）"],
        ["標籤", "否", "以逗號分隔，例：開發, 會議"],
        ["說明", "否", "任意文字"],
        ["工作分鐘數", "否", "整數（預設：0）"],
    ]
    for r_idx, row in enumerate(notes, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 6
    ws2.column_dimensions["C"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=daily_task_template.xlsx"},
    )


@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """批次匯入 Excel，回傳成功/失敗統計"""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="只接受 .xlsx 格式的 Excel 檔案")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:  # 5MB 上限
        raise HTTPException(status_code=400, detail="檔案不可超過 5MB")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="無法解析 Excel 檔案，請確認格式正確")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2))  # 跳過標題列

    if not rows:
        raise HTTPException(status_code=400, detail="Excel 中無資料列")

    if len(rows) > 500:
        raise HTTPException(status_code=400, detail="單次匯入最多 500 筆")

    created = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows, 2):
        # 略過全空列
        if all(c.value is None for c in row):
            continue

        result = _parse_row(row, row_idx)
        if isinstance(result, str):
            errors.append(result)
            continue

        task = DailyTask(
            user_id=current_user.id,
            title=result["title"],
            description=result["description"],
            status=DailyTaskStatus(result["status"]),
            progress=result["progress"],
            date=result["date"],
            work_minutes=result["work_minutes"],
        )
        db.add(task)
        await db.flush()

        for lbl in result["labels"]:
            db.add(DailyTaskLabel(daily_task_id=task.id, label=lbl))

        created += 1

    if created > 0:
        await db.commit()
    else:
        await db.rollback()

    return {
        "created": created,
        "errors": errors,
        "total_rows": len([r for r in rows if not all(c.value is None for c in r)]),
    }
